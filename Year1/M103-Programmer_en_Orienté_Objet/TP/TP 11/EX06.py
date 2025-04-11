import tkinter as tk
from tkinter import messagebox
import openpyxl

# Fonction pour enregistrer les données dans un fichier Excel
def enregistrer():
    # Récupérer les valeurs du formulaire
    nom = entry_nom.get()
    prenom = entry_prenom.get()
    age = entry_age.get()
    departement = entry_departement.get()
    email = entry_email.get()

    # Vérifier si tous les champs sont remplis
    if not nom or not prenom or not age or not departement or not email:
        messagebox.showwarning("Erreur", "Veuillez remplir tous les champs.")
        return

    # Charger ou créer un fichier Excel
    try:
        # Charger un fichier existant ou créer un nouveau
        try:
            wb = openpyxl.load_workbook("stagiaires.xlsx")
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        ws = wb.active
        # Ajouter les en-têtes de colonnes si le fichier est vide
        if ws.max_row == 1:  # Cela signifie qu'il n'y a que la ligne d'en-tête
            ws.append(["Nom", "Prénom", "Âge", "Département", "Email"])

        # Ajouter les données du stagiaire dans le fichier
        ws.append([nom, prenom, age, departement, email])

        # Sauvegarder le fichier Excel
        wb.save("stagiaires.xlsx")

        # Afficher un message de succès
        messagebox.showinfo("Succès", "Les informations du stagiaire ont été enregistrées avec succès.")

        # Réinitialiser le formulaire
        entry_nom.delete(0, tk.END)
        entry_prenom.delete(0, tk.END)
        entry_age.delete(0, tk.END)
        entry_departement.delete(0, tk.END)
        entry_email.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")

# Création de la fenêtre Tkinter
root = tk.Tk()
root.title("Enregistrement Stagiaire")
root.geometry("400x400")

# Widgets pour le formulaire
label_nom = tk.Label(root, text="Nom :", font=("Arial", 12))
entry_nom = tk.Entry(root, font=("Arial", 12))

label_prenom = tk.Label(root, text="Prénom :", font=("Arial", 12))
entry_prenom = tk.Entry(root, font=("Arial", 12))

label_age = tk.Label(root, text="Âge :", font=("Arial", 12))
entry_age = tk.Entry(root, font=("Arial", 12))

label_departement = tk.Label(root, text="Département :", font=("Arial", 12))
entry_departement = tk.Entry(root, font=("Arial", 12))

label_email = tk.Label(root, text="Email :", font=("Arial", 12))
entry_email = tk.Entry(root, font=("Arial", 12))

# Bouton pour enregistrer les données
btn_enregistrer = tk.Button(root, text="Enregistrer", command=enregistrer, font=("Arial", 12), bg="green", fg="white")

# Placement des widgets dans la fenêtre
label_nom.pack(pady=5)
entry_nom.pack(pady=5)

label_prenom.pack(pady=5)
entry_prenom.pack(pady=5)

label_age.pack(pady=5)
entry_age.pack(pady=5)

label_departement.pack(pady=5)
entry_departement.pack(pady=5)

label_email.pack(pady=5)
entry_email.pack(pady=5)

btn_enregistrer.pack(pady=20)

# Lancer l'application
root.mainloop()
